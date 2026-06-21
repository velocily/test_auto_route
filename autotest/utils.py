"""
============================================================================
结果导出模块 - 将评测结果导出到 Excel (.xlsx)
============================================================================
每题集单独输出一个 xlsx 文件，文件名格式: {模型名}_{题库名}_results.xlsx
文件内容: 题目详情 + 模型答案 + 评分 + 汇总统计
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_one_benchmark(bench_data, model_name, output_dir):
    """将一个题集的评测结果导出到一个独立的 xlsx 文件。"""
    bench_display = bench_data["benchmark_display"]
    questions = bench_data["questions"]
    results = bench_data["results"]

    is_subjective = questions[0]["type"] in ("subjective", "long_context") if questions else False
    is_long_context = questions[0]["type"] == "long_context" if questions else False

    safe_model = _safe_filename(model_name)
    safe_bench = _safe_filename(bench_display)
    filename = f"{safe_model}_{safe_bench}_results.xlsx"
    output_path = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bench_display[:31]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    high_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mid_score_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    if is_long_context:
        headers = ["题号", "类别", "问题（前200字）", "上下文（前200字）",
                   "模型回答", "得分", "评语"]
        col_widths = [8, 14, 40, 45, 55, 10, 30]
    elif is_subjective:
        headers = ["题号", "类别", "问题（前200字）", "模型回答",
                   "得分", "评语"]
        col_widths = [8, 14, 45, 60, 10, 30]
    else:
        headers = ["题号", "类别", "题目（前200字）", "正确答案",
                   "模型原始输出", "提取答案", "得分", "评语"]
        col_widths = [8, 18, 50, 14, 50, 14, 8, 30]

    max_cols = len(headers)

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f"模型: {model_name}  |  题库: {bench_display}")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border
    for col in range(2, max_cols + 1):
        ws.cell(row=1, column=col).border = thin_border
        ws.cell(row=1, column=col).fill = title_fill

    # 表头
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据行
    result_map = {r["qid"]: r for r in results}
    total_score = 0
    total_count = 0

    for row_idx, q in enumerate(questions, 3):
        qid = q["id"]
        result = result_map.get(qid, {})
        score = result.get("score", 0)
        total_score += score
        total_count += 1

        if is_long_context:
            row_data = [
                qid,
                q.get("category", ""),
                q["question_text"][:200],
                (q.get("context", "") or "")[:200],
                result.get("model_output", ""),
                score,
                result.get("comment", ""),
            ]
        elif is_subjective:
            row_data = [
                qid,
                q.get("category", ""),
                q["question_text"][:200],
                result.get("model_output", ""),
                score,
                result.get("comment", ""),
            ]
        else:
            row_data = [
                qid,
                q.get("category", ""),
                q["question_text"][:200],
                q.get("correct_answer", ""),
                result.get("model_output", ""),
                result.get("extracted_answer", ""),
                score,
                result.get("comment", ""),
            ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if is_long_context:
                score_col = 6
                extract_answer_col = 5
            elif is_subjective:
                score_col = 5
                extract_answer_col = 4
            else:
                score_col = 7
                extract_answer_col = 6

            if col_idx in (1, score_col, extract_answer_col):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if col_idx == score_col:
                if is_subjective:
                    if score >= 8:
                        cell.fill = high_score_fill
                    elif score >= 5:
                        cell.fill = mid_score_fill
                    else:
                        cell.fill = low_score_fill
                else:
                    cell.fill = high_score_fill if score == 1 else low_score_fill

    # 汇总行
    summary_row = total_count + 3

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.fill = summary_fill
        cell.border = thin_border
        cell.font = Font(bold=True, size=11)

    ws.cell(row=summary_row, column=1, value="汇总").alignment = center_align
    ws.cell(row=summary_row, column=2, value=f"共 {total_count} 题").alignment = center_align

    if is_long_context:
        avg_score = total_score / total_count if total_count > 0 else 0
        ws.cell(row=summary_row, column=3, value="").alignment = center_align
        ws.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=4)
        ws.cell(row=summary_row, column=5, value="").alignment = center_align
        ws.cell(row=summary_row, column=6,
                value=f"均分: {avg_score:.2f}/10").alignment = center_align
        ws.cell(row=summary_row, column=7,
                value=f"总分: {total_score}/{total_count * 10}").alignment = center_align
    elif is_subjective:
        avg_score = total_score / total_count if total_count > 0 else 0
        ws.cell(row=summary_row, column=3, value="").alignment = center_align
        ws.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=4)
        ws.cell(row=summary_row, column=5,
                value=f"均分: {avg_score:.2f}/10").alignment = center_align
        ws.cell(row=summary_row, column=6,
                value=f"总分: {total_score}/{total_count * 10}").alignment = center_align
    else:
        accuracy = total_score / total_count if total_count > 0 else 0
        ws.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=5)
        ws.cell(row=summary_row, column=3, value="").alignment = center_align
        ws.cell(row=summary_row, column=6, value="正确数").alignment = center_align
        ws.cell(row=summary_row, column=7,
                value=f"{total_score}/{total_count}").alignment = center_align
        ws.cell(row=summary_row, column=8,
                value=f"准确率: {accuracy:.2%}").alignment = center_align

    ws.freeze_panes = "A3"

    wb.save(output_path)
    print(f"  -> 已导出: {output_path}")
    return output_path


def export_all(all_results, model_name, output_dir):
    """将所有题集结果分别导出到独立文件。"""
    paths = []
    for bench_data in all_results:
        path = export_one_benchmark(bench_data, model_name, output_dir)
        paths.append(path)
    return paths


def export_efficiency(eff_data, model_name, output_dir):
    """导出效率测试结果到独立的 xlsx 文件。"""
    bench_display = eff_data["benchmark_display"]
    single = eff_data["single_user"]
    sweep = eff_data["concurrency_sweep"]
    stable_conc = eff_data["stable_concurrency"]
    stable_res = eff_data["stable_result"]

    safe_model = _safe_filename(model_name)
    safe_bench = _safe_filename(bench_display)
    filename = f"{safe_model}_{safe_bench}_results.xlsx"
    output_path = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 1: 概览
    ws1 = wb.active
    ws1.title = "概览"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws1.merge_cells('A1:D1')
    title_cell = ws1.cell(row=1, column=1, value=f"模型: {model_name}  |  效率测试")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border
    for col in range(2, 5):
        ws1.cell(row=1, column=col).border = thin_border
        ws1.cell(row=1, column=col).fill = title_fill

    headers = ["指标", "数值", "单位", "说明"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    rows = [
        ["单用户平均首Token延迟", f"{single.get('avg_ttft_ms', 0):.1f}", "ms", "Time To First Token"],
        ["单用户平均Token吞吐", f"{single.get('avg_throughput', 0):.2f}", "tok/s", "每秒生成token数"],
        ["单用户平均输出Token数", f"{single.get('avg_tokens', 0):.0f}", "tokens", "每轮输出token数"],
        ["稳定并行访问数量", f"{stable_conc if stable_conc else 'N/A'}", "并发", "吞吐未低于65%的最大并发"],
    ]
    if stable_res:
        rows.extend([
            ["稳定并发下平均首Token延迟", f"{stable_res.get('avg_ttft_ms', 0):.1f}", "ms", ""],
            ["稳定并发下平均Token吞吐", f"{stable_res.get('avg_throughput', 0):.2f}", "tok/s", ""],
        ])

    for row_idx, row_data in enumerate(rows, 3):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align

    for col in ['A', 'B', 'C', 'D']:
        ws1.column_dimensions[col].width = 24

    # Sheet 2: 并发扫描明细
    ws2 = wb.create_sheet(title="并发扫描明细")

    headers2 = ["并发数", "平均TTFT(ms)", "平均吞吐(tok/s)", "平均Tokens", "失败数"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, r in enumerate(sweep, 2):
        row_data = [
            r.get("concurrency", ""),
            f"{r.get('avg_ttft_ms', 0):.1f}",
            f"{r.get('avg_throughput', 0):.2f}",
            f"{r.get('avg_tokens', 0):.0f}",
            r.get("failed", 0),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            if r.get("concurrency") == stable_conc:
                cell.fill = summary_fill

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 20

    wb.save(output_path)
    print(f"  -> 已导出: {output_path}")
    return output_path


def _safe_filename(name):
    """将名称转为安全的文件名（不含非法字符）"""
    unsafe_chars = r'\/:*?"<>|'
    for c in unsafe_chars:
        name = name.replace(c, '_')
    return name
