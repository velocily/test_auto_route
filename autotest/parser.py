"""
============================================================================
题集解析模块 - 支持 MMLU（选择）、GSM8K（数学）、HellaSwag（选择）
                  职场角色测试（主观题，xlsx格式）
                  BBH语义理解（选择）、BBH数学推理（数学）
                  LongBench（长上下文主观题）
============================================================================
每种题集返回结构统一的题目列表，每题包含:
  - id, category, type, question_text, options, correct_answer, raw_answer
"""
import re
import os


def _read_file(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def parse_mmlu(filepath):
    """解析 MMLU 选择题集
    格式: Q1 [category]: question ... \n A:... B:... C:... D:... \n 答案: X
    """
    text = _read_file(filepath)
    questions = []

    blocks = re.split(r'\n(?=Q\d+\s*\[)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*\[([^\]]+)\]\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        category = m.group(2).strip()
        rest = m.group(3)

        ans_m = re.search(r'\n答案\s*:\s*(\S+)', rest)
        correct_answer = ans_m.group(1).strip() if ans_m else ""
        if ans_m:
            rest = rest[:ans_m.start()]

        opt_m = re.search(r'(?:^|\n)\s*A\s*:', rest)
        if opt_m:
            question_text = rest[:opt_m.start()].strip()
            options_text = rest[opt_m.start():].strip()
        else:
            question_text = rest.strip()
            options_text = ""

        options = _parse_options(options_text)

        questions.append({
            "id": qid,
            "category": category,
            "type": "multiple_choice",
            "question_text": question_text,
            "options": options,
            "correct_answer": correct_answer,
            "raw_answer": correct_answer,
        })

    return questions


def parse_hellaswag(filepath):
    """解析 HellaSwag 选择题集"""
    text = _read_file(filepath)
    questions = []

    blocks = re.split(r'\n(?=Q\d+\s*:)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        rest = m.group(2)

        ans_m = re.search(r'\n答案\s*:\s*(\S+)', rest)
        correct_answer = ans_m.group(1).strip() if ans_m else ""
        if ans_m:
            rest = rest[:ans_m.start()]

        opt_m = re.search(r'(?:^|\n)\s*A\s*:', rest)
        if opt_m:
            question_text = rest[:opt_m.start()].strip()
            options_text = rest[opt_m.start():].strip()
        else:
            question_text = rest.strip()
            options_text = ""

        options = _parse_options(options_text)

        questions.append({
            "id": qid,
            "category": "commonsense_reasoning",
            "type": "multiple_choice",
            "question_text": question_text,
            "options": options,
            "correct_answer": correct_answer,
            "raw_answer": correct_answer,
        })

    return questions


def parse_gsm8k(filepath):
    """解析 GSM8K 数学题集"""
    text = _read_file(filepath)
    lines = text.strip().split("\n", 1)
    if len(lines) > 1 and ("only the final answer" in lines[0] or "Do not show reasoning" in lines[0]):
        text = lines[1]

    questions = []

    blocks = re.split(r'\n(?=Q\d+\s*:)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        rest = m.group(2)

        ans_m = re.search(r'\n答案\s*:', rest)
        if ans_m:
            question_text = rest[:ans_m.start()].strip()
            answer_text = rest[ans_m.start():].strip()
        else:
            question_text = rest.strip()
            answer_text = ""

        final_answer = ""
        raw_answer = answer_text
        hash_m = re.search(r'####\s*(\S+)', answer_text)
        if hash_m:
            final_answer = hash_m.group(1).strip()

        questions.append({
            "id": qid,
            "category": "math_reasoning",
            "type": "math_fill",
            "question_text": question_text,
            "options": {},
            "correct_answer": final_answer,
            "raw_answer": raw_answer,
        })

    return questions


def _parse_options(options_text):
    """将选项文本解析为 {A: '...', B: '...', ...} 字典"""
    options = {}
    if not options_text:
        return options

    parts = re.split(r'(?:^|\n)\s*(?=[A-D]\s*:)', options_text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        m = re.match(r'([A-D])\s*:\s*(.*)', part, re.DOTALL)
        if m:
            label = m.group(1)
            content = m.group(2).strip()
            content = re.sub(r'\s+', ' ', content)
            options[label] = content

    if len(options) <= 1:
        options = {}
        single_parts = re.split(r'\s+(?=[A-D]\s*:)', options_text.strip())
        for part in single_parts:
            part = part.strip()
            if not part:
                continue
            m = re.match(r'([A-D])\s*:\s*(.*)', part, re.DOTALL)
            if m:
                label = m.group(1)
                content = m.group(2).strip()
                options[label] = content

    return options


# =========================
# 职场角色测试解析（xlsx格式）
# =========================

def parse_workplace(filepath, role="pm"):
    """解析职场角色测试 xlsx 文件"""
    import openpyxl

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_mapping = {}
    for idx, val in enumerate(header_row, 1):
        if val:
            text = str(val).strip()
            if "项目经理" in text:
                col_mapping["pm"] = idx
            elif "秘书" in text:
                col_mapping["secretary"] = idx

    if "pm" not in col_mapping:
        col_mapping["pm"] = 1
    if "secretary" not in col_mapping:
        col_mapping["secretary"] = 2

    target_col = col_mapping.get(role, 1)
    role_name = "项目经理" if role == "pm" else "秘书"

    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < target_col:
            continue
        q_text = row[target_col - 1]
        if not q_text or not str(q_text).strip():
            continue

        qid = len(questions) + 1
        questions.append({
            "id": qid,
            "category": role_name,
            "type": "subjective",
            "question_text": str(q_text).strip(),
            "options": {},
            "correct_answer": "",
            "raw_answer": "",
        })

    return questions


def parse_workplace_pm(filepath):
    """解析职场角色测试 - 项目经理题"""
    return parse_workplace(filepath, role="pm")


def parse_workplace_secretary(filepath):
    """解析职场角色测试 - 秘书题"""
    return parse_workplace(filepath, role="secretary")


# =========================
# BBH 题集解析
# =========================

def parse_bbh_semantic(filepath):
    """解析 BBH 语义理解选择题集"""
    text = _read_file(filepath)
    questions = []

    blocks = re.split(r'\n(?=Q\d+\s*:)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        rest = m.group(2)

        if qid > 10:
            break

        ans_m = re.search(r'\n答案\s*:\s*\(?([A-C])\)?', rest)
        correct_answer = ans_m.group(1).strip() if ans_m else ""
        if ans_m:
            rest = rest[:ans_m.start()]

        opt_m = re.search(r'(?:^|\n)\s*\(?\s*A\s*\)?\s*[:：]?', rest)
        if opt_m:
            question_text = rest[:opt_m.start()].strip()
            options_text = rest[opt_m.start():].strip()
        else:
            question_text = rest.strip()
            options_text = ""

        options = _parse_bbh_options(options_text)

        questions.append({
            "id": qid,
            "category": "semantic_understanding",
            "type": "multiple_choice",
            "question_text": question_text,
            "options": options,
            "correct_answer": correct_answer,
            "raw_answer": correct_answer,
        })

    return questions


def _parse_bbh_options(options_text):
    """解析 BBH 选项格式: (A) ... (B) ... (C) ..."""
    options = {}
    if not options_text:
        return options

    parts = re.split(r'\n\s*(?=\(\s*[A-C]\s*\))', options_text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        m = re.match(r'\(\s*([A-C])\s*\)\s*(.*)', part, re.DOTALL)
        if m:
            label = m.group(1)
            content = m.group(2).strip()
            options[label] = content

    return options


def parse_bbh_math(filepath):
    """解析 BBH 数学推理题集"""
    text = _read_file(filepath)
    questions = []

    blocks = re.split(r'\n(?=Q\d+\s*:)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        rest = m.group(2)

        if qid > 10:
            break

        ans_m = re.search(r'\n答案\s*:\s*(\S+)', rest)
        correct_answer = ans_m.group(1).strip() if ans_m else ""
        if ans_m:
            question_text = rest[:ans_m.start()].strip()
        else:
            question_text = rest.strip()

        question_text = question_text.rstrip("=").strip()

        questions.append({
            "id": qid,
            "category": "math_reasoning",
            "type": "math_fill",
            "question_text": question_text,
            "options": {},
            "correct_answer": correct_answer,
            "raw_answer": correct_answer,
        })

    return questions


# =========================
# LongBench 长上下文解析
# =========================

def parse_longbench(filepath):
    """解析 LongBench 长上下文题集

    支持两种格式：
    1. 新格式（每题独立 context）：
       Q1: 问题
       CONTEXT_START:
       [长文本]
       CONTEXT_END:
       答案: xxx

    2. 旧格式（共享 context，Q10 后跟长文本）：
       Q1-Q9: 短问题
       Q10: 问题 + 长文本
    """
    text = _read_file(filepath)
    questions = []

    # 检测是否为新格式（含 CONTEXT_START 标记）
    if "CONTEXT_START:" in text:
        # 新格式：每题独立 context
        blocks = re.split(r'\n(?=Q\d+\s*:)', text)
        for block in blocks:
            block = block.strip()
            if not block:
                continue

            m = re.match(r'Q(\d+)\s*:\s*(.*)', block, re.DOTALL)
            if not m:
                continue
            qid = int(m.group(1))
            rest = m.group(2)

            # 提取 context
            context = ""
            ctx_m = re.search(r'CONTEXT_START:\s*(.*?)\s*CONTEXT_END:', rest, re.DOTALL)
            if ctx_m:
                context = ctx_m.group(1).strip()
                rest = rest[:ctx_m.start()] + rest[ctx_m.end():]

            # 提取答案
            ans_m = re.search(r'\n答案\s*:\s*(.*)', rest, re.DOTALL)
            correct_answer = ""
            if ans_m:
                correct_answer = ans_m.group(1).strip()
                rest = rest[:ans_m.start()]

            question_text = rest.strip()

            questions.append({
                "id": qid,
                "category": "长上下文理解",
                "type": "long_context",
                "question_text": question_text,
                "options": {},
                "correct_answer": correct_answer,
                "raw_answer": correct_answer,
                "context": context,
            })
    else:
        # 旧格式：共享 context（向后兼容）
        long_context = ""
        m_context = re.search(r'Q10:\s*.+?\?\s*(.+?)(?=\nQ11:)', text, re.DOTALL)
        if m_context:
            long_context = m_context.group(1).strip()

        blocks = re.split(r'\n(?=Q\d+\s*:)', text)
        for block in blocks:
            block = block.strip()
            if not block:
                continue

            m = re.match(r'Q(\d+)\s*:\s*(.*)', block, re.DOTALL)
            if not m:
                continue
            qid = int(m.group(1))
            rest = m.group(2)

            if qid > 10:
                break

            if qid == 10 and long_context:
                q_text = rest.split('?')[0].strip() + '?'
            else:
                q_text = rest.strip()

            questions.append({
                "id": qid,
                "category": "长上下文理解",
                "type": "long_context",
                "question_text": q_text,
                "options": {},
                "correct_answer": "",
                "raw_answer": "",
                "context": long_context,
            })

    return questions


# =========================
# 多模态题集解析（视觉问答、图表理解、文字识别、视觉数学、多模态理解）
# =========================

def parse_multimodal(filepath):
    """解析多模态视觉题集（ChartQA / TextVQA / MathVista / VQA / MMMU 通用格式）

    格式:
        Q1 [category]: IMAGE: images/xxx.png
        问题: xxx
        A: ... B: ... C: ... D: ...
        答案: X

    每题返回结构:
        - id, category, type="visual_multiple_choice"
        - question_text, options, correct_answer, raw_answer
        - image_path (相对于题集文件所在目录)
    """
    text = _read_file(filepath)
    questions = []
    base_dir = os.path.dirname(filepath)

    blocks = re.split(r'\n(?=Q\d+\s*\[)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*\[([^\]]+)\]\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        category = m.group(2).strip()
        rest = m.group(3)

        # 提取图片路径
        image_path = ""
        img_m = re.search(r'IMAGE:\s*(\S+)', rest)
        if img_m:
            image_path = img_m.group(1).strip()
            rest = rest[:img_m.start()] + rest[img_m.end():]

        # 提取问题
        q_m = re.search(r'问题\s*:\s*(.*?)(?=\n\s*A\s*:|\n答案\s*:)', rest, re.DOTALL)
        if q_m:
            question_text = q_m.group(1).strip()
            rest_after_q = rest[q_m.end():]
        else:
            # 兜底：取 A 选项之前的内容
            opt_m = re.search(r'(?:^|\n)\s*A\s*:', rest)
            if opt_m:
                question_text = rest[:opt_m.start()].strip()
                rest_after_q = rest[opt_m.start():]
            else:
                question_text = rest.strip()
                rest_after_q = ""

        # 提取答案
        ans_m = re.search(r'\n答案\s*:\s*(\S+)', rest_after_q)
        correct_answer = ans_m.group(1).strip() if ans_m else ""
        if ans_m:
            rest_after_q = rest_after_q[:ans_m.start()]

        # 解析选项
        options = _parse_options(rest_after_q)

        questions.append({
            "id": qid,
            "category": category,
            "type": "visual_multiple_choice",
            "question_text": question_text,
            "options": options,
            "correct_answer": correct_answer,
            "raw_answer": correct_answer,
            "image_path": os.path.join(base_dir, image_path) if image_path else "",
        })

    return questions


# =========================
# 文生图（text-to-image）题集解析
# =========================

def parse_t2i(filepath):
    """解析文生图题集

    格式:
        Q1 [category]: 画一只在草地上奔跑的金毛猎犬，阳光明媚，毛发细节清晰
        评分维度: 主体准确度, 场景契合度, 画面质量
        答案: 8

    每题返回结构:
        - id, category, type="text_to_image"
        - question_text (即生图提示词), options={}, correct_answer (参考分), raw_answer
        - eval_dims (评分维度列表，用于打分模型参考)
    """
    text = _read_file(filepath)
    questions = []

    blocks = re.split(r'\n(?=Q\d+\s*\[)', text)
    for block in blocks:
        block = block.strip()
        if not block:
            continue

        m = re.match(r'Q(\d+)\s*\[([^\]]+)\]\s*:\s*(.*)', block, re.DOTALL)
        if not m:
            continue
        qid = int(m.group(1))
        category = m.group(2).strip()
        rest = m.group(3)

        # 提取评分维度
        eval_dims = []
        dims_m = re.search(r'评分维度\s*:\s*(.+?)(?=\n答案\s*:|\Z)', rest, re.DOTALL)
        if dims_m:
            dims_line = dims_m.group(1).strip()
            eval_dims = [d.strip() for d in re.split(r'[,，]', dims_line) if d.strip()]
            rest = rest[:dims_m.start()] + rest[dims_m.end():]

        # 提取参考分（用于打分模型校准，不是硬性答案）
        ans_m = re.search(r'\n答案\s*:\s*(\d+(?:\.\d+)?)', rest)
        correct_answer = ans_m.group(1).strip() if ans_m else "8"
        if ans_m:
            rest = rest[:ans_m.start()]

        # 题干即生图提示词
        question_text = rest.strip()

        questions.append({
            "id": qid,
            "category": category,
            "type": "text_to_image",
            "question_text": question_text,
            "options": {},
            "correct_answer": correct_answer,
            "raw_answer": correct_answer,
            "eval_dims": eval_dims,
        })

    return questions


# =========================
# 统一入口
# =========================
PARSER_MAP = {
    "mmlu": parse_mmlu,
    "gsm8k": parse_gsm8k,
    "hellaswag": parse_hellaswag,
    "workplace_pm": parse_workplace_pm,
    "workplace_secretary": parse_workplace_secretary,
    "bbh_semantic": parse_bbh_semantic,
    "bbh_math": parse_bbh_math,
    "longbench": parse_longbench,
    # 多模态识图题集（共用 parse_multimodal 解析器）
    "chartqa": parse_multimodal,
    "textvqa": parse_multimodal,
    "mathvista": parse_multimodal,
    "vqa": parse_multimodal,
    "mmmu": parse_multimodal,
    # 文生图题集
    "t2i": parse_t2i,
}


def parse_benchmark(benchmark_name, filepath):
    """统一解析入口，返回题目列表"""
    if benchmark_name not in PARSER_MAP:
        raise ValueError(f"不支持的题集类型: {benchmark_name}")
    return PARSER_MAP[benchmark_name](filepath)
